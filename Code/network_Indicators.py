import networkx as nx
import osmnx as ox
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import plotly.graph_objects as go
import numpy as np
from matplotlib.colors import Normalize, rgb2hex
from matplotlib import cm
import numpy as np
import scipy.stats as stats
import json
import matplotlib.path as mpltPath
from scipy.interpolate import griddata
import shapely.geometry as sg
from shapely.ops import unary_union
from scipy.ndimage import gaussian_filter
from scipy.spatial import ConvexHull
from sklearn.preprocessing import MinMaxScaler



def calculate_centrality(graph, metric, weight='length'):
        try:
            if metric == "degree":
                return dict(graph.degree())
            elif metric == "eigenvector":
                # Create undirected graph with consolidated edges
                undirected_graph = nx.Graph()
                for u, v, data in graph.edges(data=True):
                    weight = data.get('weight', 1.0)
                    if undirected_graph.has_edge(u, v):
                        undirected_graph[u][v]['weight'] += weight
                    else:
                        undirected_graph.add_edge(u, v, weight=weight)

                # Handle disconnected graphs
                if not nx.is_connected(undirected_graph):
                    lcc = max(nx.connected_components(undirected_graph), key=len)
                    undirected_graph = undirected_graph.subgraph(lcc)

                try:
                    centrality = nx.eigenvector_centrality(
                        undirected_graph, 
                        max_iter=10_000,  
                        tol=1e-9,        
                        weight='weight'
                    )
                except nx.PowerIterationFailedConvergence:
                    centrality = nx.eigenvector_centrality(
                        undirected_graph, 
                        max_iter=10_000,
                        tol=1e-3,
                        weight='weight'
                    )

                # Map back to original graph
                centrality = {node: centrality.get(node, 0.0) for node in graph.nodes()}
                max_c = max(centrality.values(), default=1)
                return {node: val / max_c for node, val in centrality.items()}
            elif metric == "pagerank":
                return nx.pagerank(graph, weight=weight)
            elif metric == "betweenness":
                return nx.betweenness_centrality(graph, weight=weight)
            elif metric == "closeness":
                # Use 'distance' parameter correctly for closeness
                if weight:
                    # For closeness, we need to invert the weight logic
                    # Create a copy of the graph with inverted weights for distance
                    distance_graph = graph.copy()
                    for u, v, d in distance_graph.edges(data=True):
                        if weight in d:
                            # Invert so larger values of weight = shorter distances
                            d[weight] = 1.0 / (d[weight] + 0.00001)  # Avoid division by zero
                    return nx.closeness_centrality(distance_graph, distance=weight)
                else:
                    print("Graph without wight")
                    return nx.closeness_centrality(graph)
            elif metric == "slc":
                return calculate_slc(graph)
            elif metric == "lsc":
                return calculate_lsc(graph, alpha=0.5)
            else:
                raise ValueError(f"Unknown metric: {metric}")
        except Exception as e:
            print(f"Error calculating centrality: {e}")
            return {node: 0.0 for node in graph.nodes()}
        
def plot_centrality(graph, metric, place_name, weight='length' ,cmap=plt.cm.jet):
    """
    Calculates specified centrality metric for a graph and saves an interactive HTML visualization. 
    Optimized for performance and large-scale network analysis.

    Parameters:
        graph (networkx.Graph): Input graph/networks
        metric (str): Centrality metric to calculate ('closeness', 'eigenvector', 'pagerank', etc.)
        place_name (str): Name of city/location for file naming
        cmap (matplotlib.colors.Colormap, optional): Colormap for visualization gradient. Default: plt.cm.viridis

    Returns:
        None
        
    The function generates and saves an interactive plotly visualization as an HTML file.
    """
    # =============================================
    # 1. Centrality Calculation with Robust Handling
    # =============================================
    

    # Calculate centrality
    centrality = calculate_centrality(graph, metric, weight=weight)

    # =============================================
    # 1.5 Normalize centrality values to [0,1] range
    # =============================================
    print("Normalizing centrality values...")
    values = list(centrality.values())
    if not values:
        print("Warning: Centrality calculation didn't produce values")
        centrality = {node: 0.0 for node in graph.nodes()}
    else:
        max_cent = max(values)
        min_cent = min(values)
        range_cent = max_cent - min_cent
        
        if range_cent > 0:
            # Normalizing to [0,1] range
            centrality = {n: (v - min_cent) / range_cent for n, v in centrality.items()}
        else:
            # Handle case where all values are the same
            centrality = {n: 0.5 for n in centrality}

    # =============================================
    # 2. Graph Preprocessing and Sampling
    # =============================================
    graph = ox.project_graph(graph)
    import random
    # Intelligent sampling for large graphs
    MAX_NODES = 100_000
    if len(graph.nodes()) > MAX_NODES:
        # Use stratified sampling to preserve graph structure
        nodes_to_keep = set(random.sample(list(graph.nodes()), MAX_NODES))
        graph = graph.subgraph(nodes_to_keep).copy()
        print(f"Large graph sampled to {MAX_NODES} nodes for performance")

    # Simplified graph processing
    try:
        graph = ox.simplify_graph(graph.copy())
        graph = ox.project_graph(graph)
    except Exception:
        pass  # Use original graph if simplification fails

    # =============================================
    # 3. Efficient Node and Position Processing
    # =============================================
    nodes = list(graph.nodes())

    # Vectorized position extraction with fallback
    try:
        positions = np.array([(graph.nodes[node]['y'], graph.nodes[node]['x']) for node in nodes])
    except KeyError:
        nodes_gdf = ox.graph_to_gdfs(graph, nodes=True, edges=False)
        positions = np.array([(geom.y, geom.x) for geom in nodes_gdf.geometry])

    # =============================================
    # 4. Centrality and Color Optimization
    # =============================================
    # Ensure centrality for all nodes with NumPy vectorization
    node_colors = np.array([centrality.get(node, 0.0) for node in nodes])

    # Apply color normalization (values should already be in [0,1] range)
    # This is only for mapping to colors, not changing the values
    #norm = Normalize(vmin=0, vmax=1)  # Fixed range now that values are normalized
    norm = Normalize(vmin=node_colors.min(), vmax=node_colors.max())

    # =============================================
    # 5. Plotly Visualization with Performance Tweaks
    # =============================================
    # Rendering parameters
    NODE_SIZE = 4  # Smaller for large graphs
    EDGE_WIDTH = 0.6  # Thin edges for performance
    
    # Use ScatterGL for massive performance improvements
    node_trace = go.Scattergl(
        x=positions[:, 1],
        y=positions[:, 0],
        mode='markers',
        marker=dict(
            size=NODE_SIZE,
            color=node_colors,
            colorscale='jet',
            colorbar=dict(
                title=f'Centrality: {metric}',
                thickness=25,
                x=1.05,
                len=0.8,
                ticksuffix='   '
            ),
            line=dict(width=0.4, color='rgba(0,0,0,0.5)'),
            opacity=0.7,
            colorbar_title=f'{metric} Centrality'
        ),
        text=[f"Node ID: {node}<br>Centrality: {centrality[node]:.6f}" for node in nodes],
        hoverinfo='text',  # Explicitly set hoverinfo to text
        hovertemplate='%{text}<extra></extra>'
    )

    # Efficient edge trace generation
    edge_x, edge_y = np.zeros(len(graph.edges()) * 3), np.zeros(len(graph.edges()) * 3)
    for i, (u, v) in enumerate(graph.edges()):
        if u in nodes and v in nodes:
            u_idx, v_idx = nodes.index(u), nodes.index(v)
            edge_x[i*3:(i+1)*3] = [positions[u_idx][1], positions[v_idx][1], None]
            edge_y[i*3:(i+1)*3] = [positions[u_idx][0], positions[v_idx][0], None]

    edge_trace = go.Scattergl(
        x=edge_x,
        y=edge_y,
        line=dict(
            width=EDGE_WIDTH,
            color='rgba(50,50,50,0.3)'
        ),
        hoverinfo='none',
        mode='lines'
    )

    # =============================================
    # 6. Layout and Export Optimization
    # =============================================
    x_range = [np.min(positions[:,1]), np.max(positions[:,1])]
    y_range = [np.min(positions[:,0]), np.max(positions[:,0])]

    layout = go.Layout(
        title=f'{metric.capitalize()} Centrality - {place_name}',
        showlegend=False,
        hovermode='closest',  # Most important setting
        hoverdistance=100,    # Distance to show hover
        spikedistance=1000,   # Spike distance
        
        margin=dict(b=20, l=20, r=120, t=40),
        xaxis=dict(
            range=x_range,
            autorange=False,
            scaleanchor="y",
            constrain='domain',
            showgrid=False, 
            zeroline=False, 
            showticklabels=False
        ),
        yaxis=dict(
            range=y_range,
            autorange=False,
            showgrid=False, 
            zeroline=False, 
            showticklabels=False
        ),
        template='plotly_white',
        height=800,
        uirevision='true'
    )

    # Create figure and configure
    fig = go.Figure(data=[edge_trace, node_trace], layout=layout)
    
    # 7. Métricas Globales de Centralidad
    # =============================================
    shannon_entropy = calculate_shannon_entropy(centrality)
    freeman_centralization = calculate_freeman_centralization(centrality)
    
    # Añadir anotación de métricas globales con estilo discreto
    fig.add_annotation(
        xref='paper', 
        yref='paper',
        x=0.02,  # Alineado a la izquierda
        y=0.98,  # Alineado arriba
        text=(
            f"• Shannon Entropy: {shannon_entropy:.4f}<br>"
            f"• Freeman Centralization: {freeman_centralization:.4f}"
        ),
        showarrow=False,
        font=dict(
            size=10, 
            color='rgba(0,0,0,0.7)'
        ),
        bgcolor='rgba(255, 255, 255, 0.5)',
        bordercolor='rgba(0,0,0,0.2)',
        borderwidth=1,
        borderpad=4,
        align='left'
    )

    config = {
        'scrollZoom': True,
        'displayModeBar': True,
        'modeBarButtonsToRemove': ['select2d', 'lasso2d'],
        'toImageButtonOptions': {
            'format': 'svg',
            'width': 1600,
            'height': 1600
        },
        'responsive': True 
    }

    # Save HTML with optimizations
    ruta_carpeta = f"Metrics_and_Graphs_Cities/{place_name}/Network_Graph"
    os.makedirs(ruta_carpeta, exist_ok=True)
    nombre_archivo = f"Centrality_{metric}_{place_name}.html"
    ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)
    
    fig.write_html(
        ruta_completa,
        config=config,
        include_plotlyjs='cdn',
        auto_open=False,
        full_html=True,
        default_width='100%',
        default_height='100%'
    )

    print(f"Optimized visualization saved in: {ruta_completa}")

def compute_edge_betweenness_data(graph, metric="betweenness", weight=None):
    """
    Calcula la centralidad de betweenness para cada edge del grafo (manteniéndolo como multigraph)
    y almacena en una lista de diccionarios, cada uno con:
      - 'u', 'v', 'k': la clave completa del edge.
      - 'value': el valor de betweenness.
      - 'x': lista con [x_u, x_v] (ya proyectados).
      - 'y': lista con [y_u, y_v] (ya proyectados).
      - 'hover': texto de hover info.
    
    Parámetros:
      - graph: Grafo de NetworkX
      - metric: Métrica a calcular (solo 'betweenness' soportada)
      - weight: Nombre del atributo de borde a usar como peso (ej. 'length', 'travel_time')
               Si es None, todos los bordes tienen peso 1.
    
    Se asume que el grafo tiene las posiciones de sus nodos en 'x' y 'y' (después de proyectar con osmnx).
    """
    if metric != "betweenness":
        raise ValueError("Solo se soporta la métrica 'betweenness' para edges.")
    
    # Calcular centralidad (las claves serán (u, v, k))
    # Pasamos el parámetro weight para usar las distancias como pesos
    edge_centrality = nx.edge_betweenness_centrality(graph, weight=weight)
    
    # Proyectar el grafo si no lo está
    try:
        if not hasattr(graph, 'simplified') or not graph.simplified:
            graph = ox.project_graph(graph)
    except Exception as e:
        print(f"Warning during projection: {e}")
    
    # Extraer posiciones usando osmnx: desempaquetamos la tupla (nodos_gdf, edges_gdf)
    nodes_gdf, edges_gdf = ox.graph_to_gdfs(graph, nodes=True, edges=True)
    # Crear un diccionario: clave = id del nodo, valor = (y, x)
    positions = {node: (geom.y, geom.x) for node, geom in nodes_gdf.geometry.items()}
    
    edge_list = []
    for (u, v, k) in graph.edges(keys=True):
        value = edge_centrality.get((u, v, k), 0.0)
        if u not in positions or v not in positions:
            continue
        
        x_u, y_u = positions[u][1], positions[u][0]
        x_v, y_v = positions[v][1], positions[v][0]
        
        # Añadir información sobre el peso usado en el hover text si está disponible
        weight_info = ""
        if weight and (u, v, k) in graph.edges:
            edge_data = graph.get_edge_data(u, v, k)
            if weight in edge_data:
                weight_info = f"<br>{weight}: {edge_data[weight]}"
        
        hover_txt = (f"Edge: {u} - {v} (key: {k})<br>"
                     f"Betweenness: {value:.4f}{weight_info}<br>"
                     f"Node {u}: (Y: {positions[u][0]:.4f}, X: {positions[u][1]:.4f})<br>"
                     f"Node {v}: (Y: {positions[v][0]:.4f}, X: {positions[v][1]:.4f})")
        
        edge_list.append({
            'u': u,
            'v': v,
            'k': k,
            'value': value,
            'x': [x_u, x_v],
            'y': [y_u, y_v],
            'hover': hover_txt
        })
    
    return edge_list

def plot_edge_centrality(edge_list, place_name, cmap=cm.jet):
    """
    Grafica los edges (sin limitar cantidad) usando la información de edge_list,
    mostrando cada edge con su color basado en el valor de betweenness.
    Se añade una barra de colores (colorbar) para ver la escala.
    Versión optimizada para mejor rendimiento.
    """
    # Calcular cmin y cmax global a partir de los valores de los edges
    values = [edge['value'] for edge in edge_list]
    cmin, cmax = (min(values), max(values)) if values else (0, 1)
    norm = Normalize(vmin=cmin, vmax=cmax)
    
    # En lugar de usar un único trace con colores personalizados,
    # agruparemos los edges por color para reducir el número de traces
    # pero mantener la capacidad de colorear por betweenness
    
    # Agrupamos edges por color
    color_groups = {}
    
    for edge in edge_list:
        value = edge['value']
        color = rgb2hex(cmap(norm(value)))
        
        if color not in color_groups:
            color_groups[color] = {
                'x': [],
                'y': [],
                'hover': [],
                'value': value  # Guardamos un valor representativo para este color
            }
        
        hover_lines = edge['hover'].split('\n')
        edge_line = hover_lines[0].strip()
        parts = edge_line.split()
        node1 = parts[1]
        node2 = parts[3]
        
        formatted_hover = f"Edge: {node2} - {node1}<br>Betweenness: {value:.4f}"
        
        # Añadimos los puntos para este edge
        color_groups[color]['x'].extend(edge['x'] + [None])
        color_groups[color]['y'].extend(edge['y'] + [None])
        color_groups[color]['hover'].extend([formatted_hover] * len(edge['x']) + [None])
    
    # Creamos un trace por cada color (mucho menos traces que antes, pero mantiene los colores)
    edge_traces = []
    for color, data in color_groups.items():
        trace = go.Scattergl(
            x=data['x'],
            y=data['y'],
            mode='lines',
            line=dict(width=1, color=color),  # Ahora color es un solo valor, no una lista
            hoverinfo='text',
            text=data['hover'],
            showlegend=False
        )
        edge_traces.append(trace)
    
    # Trazado para la barra de colores
    dummy_trace = go.Scatter(
        x=[None],
        y=[None],
        mode='markers',
        marker=dict(
            colorscale='jet',
            cmin=cmin,
            cmax=cmax,
            color=[cmin, cmax],
            colorbar=dict(title='Edge Betweenness', thickness=25, x=1.05)
        ),
        hoverinfo='none',
        showlegend=False
    )
    
    # Determinar rangos excluyendo None
    all_x = []
    all_y = []
    for edge in edge_list:
        all_x.extend([x for x in edge['x'] if x is not None])
        all_y.extend([y for y in edge['y'] if y is not None])
    
    x_range = [min(all_x), max(all_x)] if all_x else [0, 1]
    y_range = [min(all_y), max(all_y)] if all_y else [0, 1]
    
    # Configuración de layout para mejor rendimiento
    layout = go.Layout(
        title=f'Edge Betweenness Centrality - {place_name}',
        showlegend=False,
        hovermode='closest',
        margin=dict(b=20, l=20, r=20, t=40),
        xaxis=dict(
            range=x_range,
            autorange=False,
            scaleanchor="y",
            constrain='domain',
            showgrid=False,
            zeroline=False,
            showticklabels=False
        ),
        yaxis=dict(
            range=y_range,
            autorange=False,
            showgrid=False,
            zeroline=False,
            showticklabels=False
        ),
        template='plotly_white',
        height=800,
        uirevision='true'  # Ayuda a mantener el estado del zoom/pan
    )
    
    fig = go.Figure(data=edge_traces + [dummy_trace], layout=layout)
    
    # Configuración mejorada para el rendimiento
    config = {
        'scrollZoom': True,
        'displayModeBar': True,
        'modeBarButtonsToRemove': ['select2d', 'lasso2d'],
        'toImageButtonOptions': {
            'format': 'svg',
            'width': 1600,
            'height': 1600
        },
        'responsive': True
    }

    # Guardar el archivo HTML
    ruta_carpeta = f"Graphs_Cities/Graphs_for_{place_name}"
    os.makedirs(ruta_carpeta, exist_ok=True)
    nombre_archivo = f"Edge_Centrality_betweenness_{place_name}.html"
    ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)
    
    fig.write_html(
        ruta_completa,
        config=config,
        include_plotlyjs='cdn',
        auto_open=False,
        include_mathjax=False,
        full_html=True,
        default_width='100%',
        default_height='100%'
    )
    print(f"Visualización optimizada guardada en: {ruta_completa}")

def calculate_eigenvector_centrality(graph):
    # Convertir MultiDiGraph a Graph (no dirigido) consolidando aristas
    undirected_graph = nx.Graph()
    for u, v, data in graph.edges(data=True):
        weight = data.get('weight', 1.0)  # Usa 'weight' para las aristas
        if undirected_graph.has_edge(u, v):
            undirected_graph[u][v]['weight'] += weight
        else:
            undirected_graph.add_edge(u, v, weight=weight)
    
    # Verificar conectividad
    if not nx.is_connected(undirected_graph):
        # Tomar el componente conexo más grande (LCC)
        lcc = max(nx.connected_components(undirected_graph), key=len)
        undirected_graph = undirected_graph.subgraph(lcc)
    
    # Calcular eigenvector centrality con parámetros ajustados
    try:
        centrality = nx.eigenvector_centrality(
            undirected_graph, 
            max_iter=10_000,  # Aumentar iteraciones
            tol=1e-9,         # Tolerancia muy baja
            weight='weight'
        )
    except nx.PowerIterationFailedConvergence:
        # Si falla, relajar tolerancia
        centrality = nx.eigenvector_centrality(
            undirected_graph, 
            max_iter=10_000,
            tol=1e-3,
            weight='weight'
        )
    
    # Mapear al grafo original (nodos no en LCC tendrán 0)
    centrality = {node: centrality.get(node, 0.0) for node in graph.nodes()}
    
    # Normalizar para visualización
    max_c = max(centrality.values(), default=1)
    centrality = {node: val / max_c for node, val in centrality.items()}
    
    return centrality

def Numeric_coefficient_centrality(graph, metric, place_name):
    # Calcular todas las métricas
    metricas = calculate_centrality(graph, metric, weight='weight')

    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Centralidades"
    
    # Configurar cabeceras
    columnas = [
        ("Grado", 2),
        ("Intermediación", 2),
        ("Cercanía", 2),
        ("PageRank", 2),
        ("Vector Propio", 2),
        ("Centralidad semilocal",2),
        ("Centraldad local ponderada", 2)
    ]
    
    # Escribir cabecera principal
    ws.merge_cells('A1:N1')
    ws['A1'] = place_name
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Escribir cabeceras de métricas
    start_col = 1
    for nombre, span in columnas:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col+1)
        ws.cell(row=2, column=start_col, value=nombre)
        ws.cell(row=2, column=start_col).alignment = Alignment(horizontal='center')
        
        # Subcabeceras
        ws.cell(row=3, column=start_col, value="Nodo")
        ws.cell(row=3, column=start_col+1, value="Valor")
        start_col += 2
    
    # Escribir datos
    fila = 4
    for nodo in graph.nodes():
        col = 1
        for metrica in ['degree', 'betweenness', 'closeness', 'pagerank', 'eigenvector', "slc" , "lsc"]:
            try:
                valor = metricas[metrica].get(nodo, 0)
            except KeyError:
                valor = 0
                
            ws.cell(row=fila, column=col, value=nodo)
            ws.cell(row=fila, column=col+1, value=valor)
            col += 2
        fila += 1
    
    # Ajustar anchos de columna
    for idx in range(len(columnas)*2):
        ws.column_dimensions[get_column_letter(idx+1)].width = 15
    
    # Crear la carpeta si no existe
    ruta_carpeta = f"Metrics_and_Graphs_Cities/{place_name}/Metrics"
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)
    
    # Guardar archivo en la ruta especificada
    nombre_archivo = f"Centrality_{place_name}.xlsx"
    ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)
    wb.save(ruta_completa)
    
    return ruta_completa

def calculate_slc(graph):
    """
    Calcula la Centralidad Semilocal Clásica (SLC) para cada nodo en el grafo dirigido.
    
    Parámetros:
    - graph: Grafo dirigido de NetworkX.

    Retorna:
    - Un diccionario con la SLC de cada nodo.
    """
    slc = {}

    for node in graph.nodes():
        # Obtener vecinos a distancia 1
        neighbors_1 = set(graph.successors(node))  # Vecinos directos salientes

        # Calcular SLC usando la fórmula:
        slc_value = sum(len(set(graph.successors(w))) for w in neighbors_1)  # N2(w)
        slc[node] = slc_value

    # Normalizar valores entre 0 y 1
    max_slc = max(slc.values(), default=1)
    slc = {node: val / max_slc for node, val in slc.items()}

    return slc

def calculate_lsc(graph, alpha=0.5):
    """
    Calcula la Centralidad Semilocal Mejorada (LSC) para cada nodo en el grafo dirigido.
    
    Parámetros:
    - graph: Grafo dirigido de NetworkX.
    - alpha: Parámetro de ajuste (0 <= alpha <= 1).

    Retorna:
    - Un diccionario con la LSC de cada nodo.
    """
    # Obtener la centralidad SLC primero
    slc = calculate_slc(graph)
    lsc = {}

    for node in graph.nodes():
        # Obtener vecinos a distancia 1
        neighbors_1 = set(graph.successors(node))

        # Obtener vecinos a distancia 2
        neighbors_2 = set()
        for neighbor in neighbors_1:
            neighbors_2.update(set(graph.successors(neighbor)))

        # Calcular LSC usando la fórmula:
        lsc_value = sum(alpha * len(set(graph.successors(u))) + 
                        (1 - alpha) * sum(slc.get(w, 0) for w in set(graph.successors(u)))
                        for u in neighbors_1)

        lsc[node] = lsc_value

    # Normalizar valores entre 0 y 1
    max_lsc = max(lsc.values(), default=1)
    lsc = {node: val / max_lsc for node, val in lsc.items()}

    return lsc

def calculate_shannon_entropy(centrality_values):
    """
    Calcula la entropía de Shannon para valores de centralidad.
    
    Args:
        centrality_values (dict or list): Valores de centralidad de los nodos
    
    Returns:
        float: Entropía de Shannon de los valores de centralidad
    """
    # Convertir a valores de probabilidad normalizados
    values = list(centrality_values.values()) if isinstance(centrality_values, dict) else centrality_values
    probabilities = np.array(values) / np.sum(values)
    
    # Eliminar ceros para evitar problemas con log
    probabilities = probabilities[probabilities > 0]
    
    # Calcular entropía de Shannon
    entropy = -np.sum(probabilities * np.log2(probabilities))
    return entropy

def calculate_freeman_centralization(centrality_values):

    """
    Calcula el índice de centralización de Freeman para una métrica de centralidad.
    
    Args:
        centrality_values (dict or list): Valores de centralidad de los nodos
    
    Returns:
        float: Índice de centralización de Freeman
    """
    # Convertir a lista/array si es un diccionario
    values = list(centrality_values.values()) if isinstance(centrality_values, dict) else centrality_values
    
    # Encontrar el valor máximo de centralidad y calcular desviaciones
    max_centrality = np.max(values)
    n = len(values)
    
    # Suma de diferencias al máximo
    max_possible_deviation = (n - 1) * max_centrality
    actual_deviation = np.sum([max_centrality - val for val in values])
    
    # Índice de centralización de Freeman
    freeman_index = actual_deviation / max_possible_deviation if max_possible_deviation > 0 else 0
    
    return freeman_index

def plot_geo_centrality_heatmap(graph, metric, place_name, weight='length', cmap='inferno', 
                            resolution=300, log_scale=True, road_opacity=0.3, 
                            buffer_ratio=0.05, smoothing=1.0):
    """
    Creates a geographic centrality heatmap with proper city boundary and value assignment
    
    Parameters:
    -----------
    graph : networkx.Graph
        The street network graph with x, y coordinates in node attributes
    metric : str
        Centrality metric: "closeness", "betweenness", or "pagerank"
    place_name : str
        Name of the location for the title and file naming
    weight : str, default='length'
        Edge attribute to use as weight in centrality calculations
    cmap : str, default='inferno'
        Colormap for the heatmap
    resolution : int, default=300
        Resolution of the interpolation grid
    log_scale : bool, default=True
        Whether to apply log transformation to centrality values
    road_opacity : float, default=0.3
        Opacity of road network lines
    buffer_ratio : float, default=0.05
        Buffer around the road network to define the city boundary (as proportion of total extent)
    smoothing : float, default=1.0
        Smoothing factor for the interpolation
        
    Returns:
    --------
    fig : plotly.graph_objects.Figure
        The plotly figure object
    """
    
    # Calculate centrality
    centrality = calculate_centrality(graph, metric, weight=weight)
    
    # 2. Extract node coordinates and centrality values
    coords = []
    values = []
    
    for node in graph.nodes():
        if 'x' in graph.nodes[node] and 'y' in graph.nodes[node]:
            try:
                x = float(graph.nodes[node]['x'])
                y = float(graph.nodes[node]['y'])
                val = centrality.get(node, 0)
                
                if np.isfinite(x) and np.isfinite(y) and np.isfinite(val):
                    coords.append([x, y])
                    values.append(val)
            except (ValueError, TypeError):
                continue
    
    if len(coords) < 10:
        raise ValueError("Insufficient valid coordinates in graph")
    
    coords = np.array(coords)
    values = np.array(values)
    
    print(f"Extracted {len(coords)} valid nodes with coordinates")
    
    # 3. Extract the road network geometry to define city boundary
    lines = []
    for u, v in graph.edges():
        if ('x' in graph.nodes[u] and 'y' in graph.nodes[u] and 
            'x' in graph.nodes[v] and 'y' in graph.nodes[v]):
            try:
                x1, y1 = float(graph.nodes[u]['x']), float(graph.nodes[u]['y'])
                x2, y2 = float(graph.nodes[v]['x']), float(graph.nodes[v]['y'])
                
                if (np.isfinite(x1) and np.isfinite(y1) and 
                    np.isfinite(x2) and np.isfinite(y2)):
                    lines.append(sg.LineString([(x1, y1), (x2, y2)]))
            except (ValueError, TypeError):
                continue
    
    # Create a boundary from road network with buffer
    if not lines:
        raise ValueError("No valid edges found in graph")
    
    # Create a proper boundary using the road network
    road_network = unary_union(lines)
    
    # Calculate buffer distance based on the network's bounding box
    minx, miny, maxx, maxy = road_network.bounds
    width = maxx - minx
    height = maxy - miny
    buffer_distance = max(width, height) * buffer_ratio
    
    # Create buffered boundary
    boundary = road_network.buffer(buffer_distance)
    
    # Ensure boundary is valid and simplify for performance
    if not boundary.is_valid:
        boundary = boundary.buffer(0)  # Fix invalid geometries
    boundary = boundary.simplify(buffer_distance/10)
    
    print(f"Created city boundary with buffer distance: {buffer_distance:.6f}")
    
    # 4. Normalize centrality values (before interpolation)
    if len(values) > 0:
        if log_scale and np.min(values) >= 0:
            # Add small constant to avoid log(0)
            min_positive = np.min(values[values > 0]) if np.any(values > 0) else 1e-6
            values = np.log1p(values + min_positive/10)
        
        # Normalize to [0,1] range
        values = (values - np.min(values)) / (np.max(values) - np.min(values) + 1e-10)
    
    # 5. Create proper interpolation grid that covers the boundary
    bounds = boundary.bounds
    x_min, y_min, x_max, y_max = bounds
    
    # Add small padding
    padding = 0.01
    width = x_max - x_min
    height = y_max - y_min
    
    x_min -= width * padding
    x_max += width * padding
    y_min -= height * padding
    y_max += height * padding
    
    # Create grid
    x_grid = np.linspace(x_min, x_max, resolution)
    y_grid = np.linspace(y_min, y_max, resolution)
    xx, yy = np.meshgrid(x_grid, y_grid)
    
    # 6. Perform correct interpolation using natural neighbor or linear interpolation
    grid_points = np.vstack([xx.ravel(), yy.ravel()]).T
    
    print("Performing interpolation...")
    try:
        # Use natural neighbor interpolation for better results
        z = griddata(coords, values, grid_points, method='linear', fill_value=np.nan)
        z = z.reshape(xx.shape)
        
        # Apply Gaussian smoothing for visual appeal, adjustable via parameter
        if smoothing > 0:
            z = gaussian_filter(z, sigma=smoothing)
    except Exception as e:
        print(f"Interpolation error: {str(e)}")
        # Create empty grid as fallback
        z = np.zeros(xx.shape) * np.nan
    
    # 7. Mask points outside city boundary
    print("Masking areas outside city boundary...")
    mask = np.zeros_like(xx, dtype=bool)
    
    # Convert boundary to Path for efficient point-in-polygon tests
    if isinstance(boundary, sg.Polygon):
        exterior = list(boundary.exterior.coords)
        path = mpltPath.Path(exterior)
        mask = path.contains_points(grid_points).reshape(xx.shape)
    elif isinstance(boundary, sg.MultiPolygon):
        # Handle multiple polygons
        for poly in boundary.geoms:
            exterior = list(poly.exterior.coords)
            path = mpltPath.Path(exterior)
            mask = mask | path.contains_points(grid_points).reshape(xx.shape)
    
    # Apply mask
    z = np.where(mask, z, np.nan)
    
    # 8. Create the heatmap visualization
    print("Creating visualization...")
    fig = go.Figure()
    
    # Add the heatmap trace
    fig.add_trace(go.Heatmap(
        x=x_grid,
        y=y_grid,
        z=z,
        colorscale=cmap,
        zsmooth='best',
        hoverinfo='none',
        colorbar=dict(
            title=f'{metric.capitalize()} Centrality',
            thickness=25,
            len=0.75,
            title_font=dict(size=16, color='white'),
            tickfont=dict(color='white')
        ),
        zauto=True
    ))
    
    # Add road network overlay
    edge_x = []
    edge_y = []
    
    for line in lines:
        coords = list(line.coords)
        for i in range(len(coords) - 1):
            edge_x.extend([coords[i][0], coords[i+1][0], None])
            edge_y.extend([coords[i][1], coords[i+1][1], None])
    
    if edge_x:
        fig.add_trace(go.Scattergl(
            x=edge_x,
            y=edge_y,
            mode='lines',
            line=dict(color=f'rgba(255,255,255,{road_opacity})', width=1.0),
            hoverinfo='none',
            showlegend=False
        ))
    
    # 9. Set layout with proper sizing and aspect ratio
    # Calculate aspect ratio based on boundary dimensions
    
    fig.update_layout(
        title={
            'text': f'{metric.capitalize()} Centrality - {place_name}',
            'font': {'size': 24, 'color': 'white'},
            'x': 0.5,
            'xanchor': 'center',
            'y': 0.98
        },
        margin=dict(l=10, r=10, t=80, b=10),
        paper_bgcolor='rgba(0,0,0,1)',
        plot_bgcolor='rgba(0,0,0,1)',
        xaxis=dict(
            showgrid=False, 
            zeroline=False, 
            showticklabels=False,
            range=[x_min, x_max]
        ),
        yaxis=dict(
            showgrid=False, 
            zeroline=False, 
            showticklabels=False, 
            scaleanchor="x",
            scaleratio=1,
            range=[y_min, y_max]
        )
    )
    
   # Save results
   # Save results
    output_dir = f"Metrics_and_Graphs_Cities/{place_name}/Heatmap_Graph"
    os.makedirs(output_dir, exist_ok=True)

    try:
        filename = f"{output_dir}/Heatmap_{metric}_{place_name.replace(' ', '_').replace(',', '')}.html"
        
        # Increase figure size by 25% more than previous solution
        fig.update_layout(
            width=1800,  # Increased from 1200
            height=900,  # Increased from 800
            autosize=False
        )
        
        # Keep the original Plotly HTML generation
        html_string = fig.to_html(
            include_plotlyjs="cdn",
            full_html=False,
            config={
                'displayModeBar': True,
                'responsive': True,
                'toImageButtonOptions': {
                    'format': 'svg',
                    'filename': f"{metric}_{place_name}",
                }
            }
        )
        
        # Modified HTML template with proper centering
        with open(filename, 'w') as f:
            f.write(f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>{metric.capitalize()} Centrality - {place_name}</title>
                <style>
                    html, body {{
                        background-color: black;
                        margin: 0;
                        padding: 0;
                        width: 100%;
                        height: 100%;
                        color: white;
                        overflow: hidden;
                    }}
                    #graph-container {{
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        width: 100%;
                        height: 100%;
                    }}
                    .js-plotly-plot {{
                        margin: 0 auto;
                    }}
                </style>
            </head>
            <body>
                <div id="graph-container">
                    {html_string}
                </div>
            </body>
            </html>
            """)
        
        print(f"Saved to {filename}")
    except Exception as e:
        print(f"Error saving file: {str(e)}")
    return fig