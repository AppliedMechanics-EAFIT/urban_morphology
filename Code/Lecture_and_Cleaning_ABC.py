## Functions to read the xlsx format of mobility data
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
import pandas as pd
import geopandas as gpd
from matplotlib.colors import ListedColormap
from Polygon_clustering import load_polygon_stats_from_txt, classify_polygon

## Class to define the different mobility information (The ABC of mobility)
class CiudadesABC:
    def __init__(self, index, ObsID, year, LastObservation, City, metro_names, Country, continent, 
                 region, subregion, state_name, state_abbr, population, longitude, latitude, 
                 Walking, Cycling, Motorbikes, Active, Bus, Car, IncomeGroup, DataSource, DataLink, GDPPP_2022):
        self.index = index  # Format like M10001
        self.ObsID = ObsID  # Format like ID0005
        self.year = year  # Year in numeric format
        self.LastObservation = LastObservation  # Variable text
        self.City = City  # City name
        self.metro_names = metro_names  # Metropolitan area name
        self.Country = Country  # Country
        self.continent = continent  # Continent
        self.region = region  # Region
        self.subregion = subregion  # Subregion
        self.state_name = str(state_name) if state_name not in [None, "", " "] else "Not applicable" # State name
        self.state_abbr = str(state_abbr) if state_abbr not in [None, "", " "] else "Not applicable"  # State abbreviation
        self.population = int(population) if population else None  # Population
        self.longitude = float(longitude)  # Geographic longitude
        self.latitude = float(latitude)  # Geographic latitude
        self.Walking = float(Walking) if Walking != "NA" else None  # Percentage of walking
        self.Cycling = float(Cycling) if Cycling != "NA" else None  # Percentage of cycling
        self.Motorbikes = float(Motorbikes) if Motorbikes != "NA" else None  # Percentage of motorbike use
        self.Active = float(Active)  # Percentage of active mobility
        self.Bus = float(Bus)  # Percentage of bus use
        self.Car = float(Car)  # Percentage of car use
        self.IncomeGroup = IncomeGroup  # Income group
        self.DataSource = DataSource  # Data source
        self.DataLink = DataLink  # Data link
        self.GDPPP_2022 = float(GDPPP_2022)  # Adjusted GDP per capita in 2022

    def __repr__(self):
        return (
            f"CiudadesABC(index={self.index}, ObsID={self.ObsID}, year={self.year}, "
            f"LastObservation={self.LastObservation}, City={self.City}, "
            f"metro_names={self.metro_names}, Country={self.Country}, continent={self.continent}, "
            f"region={self.region}, subregion={self.subregion}, state_name={self.state_name}, "
            f"state_abbr={self.state_abbr}, population={self.population}, longitude={self.longitude}, "
            f"latitude={self.latitude}, Walking={self.Walking}, Cycling={self.Cycling}, "
            f"Motorbikes={self.Motorbikes}, Active={self.Active}, Bus={self.Bus}, Car={self.Car}, "
            f"IncomeGroup={self.IncomeGroup}, DataSource={self.DataSource}, DataLink={self.DataLink}, "
            f"GDPPP_2022={self.GDPPP_2022})"
        )

## Function to read the xlsx file with the information
def read_nodes_from_excel(filename, sheet_name):
    wb = load_workbook(filename=filename, data_only=True)
    sheet = wb[sheet_name]

    nodes = []

    # Iterate over rows starting from the second (excluding headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        node = CiudadesABC(
            index=row[0],
            ObsID=row[1],
            year=row[2],
            LastObservation=row[3],
            City=row[4],
            metro_names=row[5],
            Country=row[6],
            continent=row[7],
            region=row[8],
            subregion=row[9],
            state_name=row[10],
            state_abbr=row[11],
            population=row[12],
            longitude=row[13],
            latitude=row[14],
            Walking=row[15],
            Cycling=row[16],
            Motorbikes=row[17],
            Active=row[18],
            Bus=row[19],
            Car=row[20],
            IncomeGroup=row[21],
            DataSource=row[22],
            DataLink=row[23],
            GDPPP_2022=row[24],
        )
        nodes.append(node)

    wb.close()
    return nodes

def clean_and_filter_data(file_path, output_file_ABC, report_file_ABC):
    # Load the Excel file
    df = pd.read_excel(file_path, dtype=str)  # Read as string to avoid type errors

    # ---- Remove duplicates keeping the most updated version ----
    key_columns = df.columns[4:10]  # Columns E to J for comparison
    removed_rows = pd.DataFrame()  # DataFrame to store removed rows

    # Identify duplicates based on key columns
    duplicates = df.duplicated(subset=key_columns, keep=False)  # Mark both duplicates as `True`

    # Filter cases where duplicates exist and one of them has "YES" in column D
    for _, group in df[duplicates].groupby(list(key_columns)):
        if "YES" in group.iloc[:, 3].values:  # Column D is at index 3
            to_remove = group[group.iloc[:, 3] == "NO"]  # Keep "YES", remove "NO"
            removed_rows = pd.concat([removed_rows, to_remove])
            df = df.drop(to_remove.index)

    # ---- Remove rows with 3 or more missing values in columns P to U ----
    key_columns_missing = df.columns[15:21]  # Columns P to U (indices 15 to 20 in 0-based indexing)
    missing_values_rows = df[df[key_columns_missing].isna().sum(axis=1) >= 3]  # Filter rows with 3+ missing values

    # Remove those rows from the original DataFrame and add them to the report
    removed_rows = pd.concat([removed_rows, missing_values_rows])
    df = df.drop(missing_values_rows.index)

    # Save the cleaned data
    df.to_excel(output_file_ABC, index=False)
    
    # Save the removed rows in a report
    removed_rows.to_excel(report_file_ABC, index=False)

    print(f"Total removed rows: {len(removed_rows)}")

def process_mobility_data_for_medellin(area_type='urban_and_rural'):
    # Set paths and options according to the area type
    if area_type == 'urban':
        a_path = "Poligonos_Medellin/Json_files/EOD_2017_SIT_only_AMVA_URBANO.geojson"
        include_absolute = False
        output_xlsx = "Poligonos_Medellin/Resultados/Statics_Results/URBAN/Poligonos_Clasificados_Movilidad_URBANO.xlsx"
        output_image = "Poligonos_Medellin/Resultados/Statics_Results/URBAN/map_poligonosA_urbano_classified.png"
    else:
        a_path = "Poligonos_Medellin/EOD_2017_SIT_only_AMVA.shp"
        include_absolute = False
        output_xlsx = "Poligonos_Medellin/Resultados/Statics_Results/URBAN_AND_RURAL/Poligonos_Clasificados_Movilidad_Urban_and_Rural.xlsx"
        output_image = "Poligonos_Medellin/Resultados/Statics_Results/URBAN_AND_RURAL/map_poligonosA_classified_Urban_and_Rural.png"
    
    # Common paths
    stats_txt = "Poligonos_Medellin/Resultados/poligonos_stats_ordenado.txt"
    matches_csv = "Poligonos_Medellin/Resultados/Matchs_A_B/matches_by_area.csv"
    shpB = "Poligonos_Medellin/eod_gen_trips_mode.shp"
    
    # 1) Load statistics
    stats_dict = load_polygon_stats_from_txt(stats_txt)
    print(f"Loaded stats for {len(stats_dict)} polygons (sub-polygons).")
    
    # 2) Load matches
    df_matches = pd.read_csv(matches_csv)
    print("df_matches sample:\n", df_matches.head(), "\n")
    
    # 3) Read shapefile B (mobility)
    gdfB = gpd.read_file(shpB)
    print("Columns B:", gdfB.columns)
    
    # 4) Read GeoDataFrame A (geometry)
    gdfA = gpd.read_file(a_path)
    print(f"Read {len(gdfA)} polygons in {'GeoJSON URBANO' if area_type == 'urban' else 'SHP'}.")
    
    # 5) Build final DataFrame
    final_rows = []
    for _, row in df_matches.iterrows():
        idxA = row["indexA"]
        idxB = row["indexB"]
        ratio = row["area_ratio"]
        
        # Get stats and classify pattern
        key_stats = (idxA, 0)
        poly_stats = stats_dict.get(key_stats, {})
        pattern = classify_polygon(poly_stats)
        
        # Extract mobility data
        rowB = gdfB.loc[idxB]
        p_walk_ = rowB.get("p_walk", 0)
        p_tpc_ = rowB.get("p_tpc", 0)
        p_sitva_ = rowB.get("p_sitva", 0)
        p_auto_ = rowB.get("p_auto", 0)
        p_moto_ = rowB.get("p_moto", 0)
        p_taxi_ = rowB.get("p_taxi", 0)
        p_bike_ = rowB.get("p_bike", 0)
        
        # Base data for the row
        row_data = {
            "indexA": idxA,
            "indexB": idxB,
            "area_ratio": ratio,
            "street_pattern": pattern,
            "p_walk": p_walk_,
            "p_tpc": p_tpc_,
            "p_sitva": p_sitva_,
            "p_auto": p_auto_,
            "p_moto": p_moto_,
            "p_taxi": p_taxi_,
            "p_bike": p_bike_
        }
        
        # Add absolute values if needed
        if include_absolute:
            row_data.update({
                "Auto": rowB.get("Auto", 0),
                "Moto": rowB.get("Moto", 0),
                "Taxi": rowB.get("Taxi", 0)
            })
        
        final_rows.append(row_data)
    
    # Define columns for the final DataFrame
    columns = ["indexA", "indexB", "area_ratio", "street_pattern"]
    if include_absolute:
        columns += ["Auto", "Moto", "Taxi"]
    columns += ["p_walk", "p_tpc", "p_sitva", "p_auto", "p_moto", "p_taxi", "p_bike"]
    
    df_final = pd.DataFrame(final_rows)[columns]
    
    # Save Excel
    df_final.to_excel(output_xlsx, index=False)
    print(f"Final Excel saved at {output_xlsx} with {len(df_final)} rows.\n")
    
    # 6) Plot classified polygons
    gdfA["pattern"] = gdfA.index.map(df_final.set_index("indexA")["street_pattern"])
    
    # Pattern-to-color mapping
    color_mapping = {
        'gridiron': 'Green',
        'cul_de_sac': 'Red',
        'hibrido': 'Blue',
        'organico': 'Yellow'
    }
    
    # Convert to ordered categories
    categories = ['gridiron', 'cul_de_sac', 'hibrido', 'organico']
    gdfA['pattern'] = pd.Categorical(gdfA['pattern'], categories=categories)
    
    # Create custom colormap
    cmap = ListedColormap([color_mapping[cat] for cat in categories])
    
    fig, ax = plt.subplots(figsize=(10, 10))
    
    # Plot with black borders
    gdfA.plot(
        column="pattern",
        ax=ax,
        legend=True,
        cmap=cmap,
        edgecolor='black',
        linewidth=1,
    )
    
    # Customize legend
    legend = ax.get_legend()
    legend.set_title('Street Pattern')
    legend.set_bbox_to_anchor((1.05, 1))

    title = "Classified Polygons A" if area_type != 'urban' else "Urban GeoJSON - Classified Polygons"
    gdfA.plot(column="pattern", ax=ax, legend=True, cmap="Set2")
    ax.set_title(title)
    plt.tight_layout()
    plt.savefig(output_image, dpi=300)
    print(f"Map saved at {output_image}")

# Usage example
if __name__ == "__main__":
    # Process urban and rural data
    process_mobility_data_for_medellin(area_type='urban_and_rural')
    
    # Process only urban data
    process_mobility_data_for_medellin(area_type='urban')
